# ===== app/parsing/xlsx_parser.py =====
import time
import openpyxl
from app.parsing.base import BaseParser
from app.parsing.schema import ParsedDocument, PageInfo, ParsedTable, TableCell

class XlsxParser(BaseParser):
    name = "xlsx_parser"
    version = "1.0"

    async def parse(self, document_id: str, content: bytes) -> ParsedDocument:
        import io
        t0 = time.perf_counter()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        tables, raw_text_parts, pages = [], [], []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            pages.append(PageInfo(page_number=sheet_idx))
            cells = []
            for r_idx, row in enumerate(ws.iter_rows()):
                for c_idx, cell in enumerate(row):
                    if cell.value is not None:
                        cells.append(TableCell(row=r_idx, col=c_idx, text=str(cell.value)))
                        raw_text_parts.append(str(cell.value))
            tables.append(ParsedTable(table_id=f"{document_id}_{sheet_name}", page=sheet_idx, order=sheet_idx, cells=cells))

        return ParsedDocument(
            document_id=document_id, source_format="xlsx", parser_name=self.name, parser_version=self.version,
            pages=pages, blocks=[], tables=tables, images=[], raw_text=" ".join(raw_text_parts),
            reading_order_applied=False, ocr_used=False, processing_time_ms=(time.perf_counter() - t0) * 1000,
        )
